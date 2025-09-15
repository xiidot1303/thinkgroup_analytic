from django.contrib import admin
from app.models import Branch, Category, Expense
from rangefilter.filters import DateRangeFilter
from app.forms import ExpenseForm
from unfold.admin import ModelAdmin as UnfoldModelAdmin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from unfold.decorators import action
from unfold.enums import ActionVariant
from unfold.contrib.filters.admin import RangeDateFilter, RangeDateTimeFilter
from django.http import HttpRequest
from django.contrib.admin.views.main import ChangeList
import urllib


@admin.register(Branch)
class BranchAdmin(UnfoldModelAdmin):
    list_display = ('name', 'address')
    search_fields = ('name', 'address') 

@admin.register(Category)
class CategoryAdmin(UnfoldModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

@admin.register(Expense)
class ExpenseAdmin(UnfoldModelAdmin):
    # form 
    list_display = ('branch', 'category', 'amount', 'staff', 'date')
    list_filter_submit = True
    list_filter = ('branch', 'category', ("date", RangeDateFilter))
    search_fields = ('branch__name', 'category__name', 'description')

    
    actions_list = ["download_report"]

    @action(description="📊 Download report (by categories)", variant=ActionVariant.PRIMARY)
    def download_report(self, request: HttpRequest):
        referer = request.META.get("HTTP_REFERER", "")
        if "?" in referer:
            query = referer.split("?", 1)[1]
            params = urllib.parse.parse_qs(query)

        # --- Clone request with parsed GET params ---
        mutable_get = request.GET.copy()
        for k, v in params.items():
            mutable_get.setlist(k, v)
        request.GET = mutable_get
        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses by Categories"

        # Headers
        ws.append(["Kategoriya", "Xarajatlar summasi (so‘m)"])

        # Group by category
        category_totals = {}
        for exp in queryset:
            category_totals.setdefault(exp.category.name, 0)
            category_totals[exp.category.name] += exp.amount

        # Fill data
        for category, total in category_totals.items():
            ws.append([category, total])

        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="expenses_report.xlsx"'
        wb.save(response)
        return response